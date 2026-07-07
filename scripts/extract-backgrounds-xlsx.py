#!/usr/bin/env python3
"""Extract enriched backgrounds from 个性与背景创建规则.xlsx."""
import json
import os
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "个性与背景创建规则.xlsx")
OUT = os.path.join(ROOT, "advisor", "chargen", "_backgrounds_xlsx_raw.json")

LABEL_MAP = {
    "背景名称": "name",
    "简介": "intro",
    "生命值加成": "hpBonus",
    "基础熟练项": "baseSkills",
    "专业熟练项": "profSkills",
    "其他": "other",
    "资金": "funds",
    "装备": "equipment",
    "特性": "traitName",
    "特性描述": "traitDesc",
    "神祇信仰": "deity",
    "地点": "location",
    "性格": "personality",
    "羁绊": "bonds",
    "缺点": "flaws",
}


def parse_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    data = {}
    for r in rows:
        if not r or len(r) < 2:
            continue
        key = r[1]
        if key is None:
            continue
        key = str(key).strip()
        if key not in LABEL_MAP:
            continue
        val = r[2] if len(r) > 2 else None
        if val is None:
            continue
        val = str(val).strip()
        if val:
            data[LABEL_MAP[key]] = val
    if "name" not in data:
        data["name"] = ws.title
    if data.get("equipment"):
        data["equipmentList"] = [
            x.strip() for x in data["equipment"].replace("\n", "、").split("、") if x.strip()
        ]
    lore_parts = []
    for k in ("intro", "location", "personality", "bonds", "flaws", "traitDesc"):
        if data.get(k):
            lore_parts.append(data[k])
    data["lore"] = "\n\n".join(lore_parts)
    return data


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    backgrounds = []
    for sn in wb.sheetnames:
        bg = parse_sheet(wb[sn])
        if bg:
            backgrounds.append(bg)
    wb.close()
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump({"count": len(backgrounds), "backgrounds": backgrounds}, f, ensure_ascii=False, indent=2)
    print(len(backgrounds))


if __name__ == "__main__":
    main()
