# -*- coding: utf-8 -*-
"""Build 斯诺德物资大全.xlsx -> 斯诺德跑团/store_data.js

修复（v1.0.7227）：
- 过滤表头泄漏行（名称 == '名称' 的重复表头行）
- 载重列错位修复：组内取第一个含「磅」的候选（动物/载具/酒水多列错位时取真重量；
  无磅候选视为源数据错误，置空并输出警告）
- 小类（类别列）写入输出 cat 字段；去掉换行符；'类别' 伪小类不污染后续行
"""
import io
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

SRC = r'D:\Download\scholar-agent-main\斯诺德物资大全.xlsx'
OUT = r'D:\Download\scholar-agent-main\斯诺德跑团\store_data.js'

HEADER_NAMES = {'类别': 'cat', '名称': 'name', '售价': 'price', '载重': 'weight', '负重': 'weight',
                '简介': 'desc', '品质': 'quality'}

SHEET_CATS = {1: '杂物', 2: '武器', 3: '草药', 4: '宝石', 5: '零件',
              6: '常见道具', 7: '卷轴', 8: '生物素材', 9: '魔法道具'}

# 源数据人工修正（xlsx 错误无法程序化推导的值）
OVERRIDES = {
    '虾肉': {'weight': '1磅'},  # xlsx 载重列为 '2银币'（笔误），真值 1磅
}

wb = openpyxl.load_workbook(SRC, data_only=True)
store = {}
order = []
total = 0
skipped_header_rows = 0
warn_weight = []

for ws in wb.worksheets:
    max_col = ws.max_column
    sheet_cat = SHEET_CATS.get(wb.worksheets.index(ws) + 1, '未分类')
    # 表头行（第 2 行）
    header = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=2, column=c).value
        if v and str(v).strip() in HEADER_NAMES:
            header[c] = HEADER_NAMES[str(v).strip()]
    if not header:
        continue
    # 数据行（第 3 行起）
    # 小类按「列组」作用域维护（物件 sheet 有 5 个并行列组，各自独立的类别列）
    cat_scope = {}
    group = None
    def flush():
        global total
        if group and group.get('name'):
            cat = sheet_cat
            if cat not in store:
                store[cat] = []
                order.append(cat)
            for _prev in store[cat]:
                if _prev['name'] == group['name']:
                    return
            _ov = OVERRIDES.get(group['name'])
            if _ov:
                for _ok, _ovv in _ov.items():
                    group[_ok] = _ovv
            store[cat].append(group)
            total += 1
    for r in range(3, ws.max_row + 1):
        row_vals = {}
        for c, key in header.items():
            row_vals[c] = ws.cell(row=r, column=c).value
        # 按组切分：'类别' 列是组起点
        cols = sorted(header.keys())
        starts = [c for c in cols if header[c] == 'cat']
        groups_bound = []
        for i, s in enumerate(starts):
            end = starts[i + 1] if i + 1 < len(starts) else (max(cols) + 1)
            groups_bound.append((s, end))
        for (s, end) in groups_bound:
            cat_val = row_vals.get(s)
            if cat_val is not None and str(cat_val).strip():
                flush()
                new_cat = str(cat_val).strip().replace('\n', '')
                if new_cat != '类别':
                    # 表头泄漏行的 '类别' 标记不更新作用域，防止污染后续行
                    cat_scope[s] = new_cat
                group = {'cat': cat_scope.get(s, '') or '', '_scope': cat_scope.get(s, '') or ''}
            elif group is None:
                group = {'cat': cat_scope.get(s, '') or '', '_scope': cat_scope.get(s, '') or ''}
            name = None
            price = None
            desc = None
            quality = None
            weight_cands = []
            for c in range(s, end):
                key = header.get(c)
                if key is None:
                    continue
                v = row_vals.get(c)
                if v is None:
                    continue
                sv = str(v).strip()
                if key == 'name' and not name and sv:
                    name = sv
                elif key == 'price' and not price and sv:
                    price = sv
                elif key == 'weight':
                    if sv:
                        weight_cands.append(sv)
                elif key == 'desc' and not desc and sv:
                    desc = sv
                elif key == 'quality' and not quality and sv:
                    quality = sv
            # 载重：取第一个含「磅」的候选（动物/载具/酒水错位列取真重量）
            weight = None
            if weight_cands:
                for cand in weight_cands:
                    if '磅' in cand:
                        weight = cand
                        break
                if weight is None:
                    warn_weight.append((ws.title, r, name, weight_cands))
                    weight = ''  # 源数据错误：货币值当载重，置空
            if name:
                if name == '名称':
                    # 表头泄漏行（sheet 内的重复表头）→ 跳过
                    skipped_header_rows += 1
                    group = None
                    continue
                group['cat'] = group.get('_scope', '')
                group.pop('_scope', None)
                group['name'] = name
                if price:
                    group['price'] = price
                if weight:
                    group['weight'] = weight
                if desc:
                    group['desc'] = desc
                if quality:
                    group['quality'] = quality
                flush()
                group = None
    flush()

# 输出 JS
buf = io.StringIO()
buf.write('// Auto-generated from 斯诺德物资大全.xlsx — store data\n')
buf.write('var STORE_DATA = {\n')
for cat in order:
    items = store[cat]
    buf.write('  %s: [\n' % repr(cat))
    for it in items:
        buf.write('    {name:%s' % repr(it['name']))
        buf.write(',cat:%s' % repr(it.get('cat', '')))
        if 'price' in it:
            buf.write(',price:%s' % repr(it['price']))
        if 'weight' in it:
            buf.write(',weight:%s' % repr(it['weight']))
        if 'desc' in it:
            buf.write(',desc:%s' % repr(it['desc']))
        if 'quality' in it:
            buf.write(',quality:%s' % repr(it['quality']))
        buf.write('},\n')
    buf.write('  ],\n')
buf.write('};\n')
io.open(OUT, 'w', encoding='utf-8').write(buf.getvalue())
print('categories:', order)
print('total items:', total)
print('skipped header rows:', skipped_header_rows)
print('weight warnings:', warn_weight)
print('written:', OUT)
