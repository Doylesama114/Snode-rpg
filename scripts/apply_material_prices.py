#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
奇械师 / 艺术与创造之神 图纸参考价格补齐。

规则：参考价格 = 材料成本加算（物资大全单价，按份/磅/米计价）；
      缺失材料按 “X金币Y银币+材料名1份+材料名2份” 标注；
      无材料（疫苗射线/全感知乐谱）列为“待定”。

用法：
  python scripts/apply_material_prices.py --dry     # 生成清单不写入
  python scripts/apply_material_prices.py --apply   # 写 JSON + 重建页面
"""
from __future__ import annotations

import json
import re
import shutil
import subprocess
import sys
from pathlib import Path

import docx
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

from apply_class_extract import extract_to_block  # noqa: E402
from class_sync_core import (  # noqa: E402
    append_tables_to_search,
    build_data_search,
    build_detail_html,
    build_skill_data_attrs,
    json_to_fx_entry,
    patch_html,
    tier_label_from_skill,
)


# ---------------------------------------------------------------------------
# 价格字典
# ---------------------------------------------------------------------------
def parse_price_copper(s):
    s = (s or "").strip()
    if not s or s == "-":
        return None
    g = re.search(r"(\d+(?:\.\d+)?)金币", s)
    y = re.search(r"(\d+(?:\.\d+)?)银币", s)
    c = re.search(r"(\d+(?:\.\d+)?)铜币", s)
    total = 0
    if g:
        total += int(round(float(g.group(1)) * 100))
    if y:
        total += int(round(float(y.group(1)) * 10))
    if c:
        total += int(round(float(c.group(1))))
    return total if (g or y or c) else None


def parse_weight_lb(s):
    m = re.search(r"(\d+(?:\.\d+)?)磅", s or "")
    return float(m.group(1)) if m else None


def build_price_dict():
    """从物资大全 xlsx 提取 名称 -> {copper, weight_lb, per_lb}。"""
    wb = openpyxl.load_workbook(
        str(ROOT / "斯诺德物资大全.xlsx"), read_only=True, data_only=True
    )
    D: dict[str, dict] = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        for r in rows:
            if not any(str(v).strip() == "名称" for v in r if v is not None):
                continue
            if not any(str(v).strip() == "售价" for v in r if v is not None):
                continue
            header = [str(v).strip() if v is not None else "" for v in r]
            for seg in range(0, len(header), 6):
                lo, hi = seg, min(seg + 6, len(header))
                nc = pc = wc = None
                for i in range(lo, hi):
                    if header[i] == "名称" and nc is None:
                        nc = i
                    if header[i] == "售价" and pc is None:
                        pc = i
                    if header[i] == "载重" and wc is None:
                        wc = i
                if nc is None or pc is None:
                    continue
                for row in rows[rows.index(r) + 1 :]:
                    if len(row) <= max(nc, pc):
                        continue
                    nm = str(row[nc]).strip() if row[nc] is not None else ""
                    pv = row[pc] if pc < len(row) else None
                    if not nm or pv is None:
                        continue
                    p = parse_price_copper(str(pv))
                    if p is None:
                        continue
                    wt = (
                        parse_weight_lb(str(row[wc]))
                        if wc is not None and wc < len(row) and row[wc] is not None
                        else None
                    )
                    D.setdefault(
                        nm,
                        {
                            "copper": p,
                            "weight_lb": wt,
                            "per_lb": round(p / wt, 6) if wt else None,
                        },
                    )
            break
    return D


# 官方参考价倒推的基础物品价（铜币），物资大全无该条目；
# 值可以是 int（按份/个/瓶）或 dict（按磅等特殊计价）。
EXTRA_PRICE = {
    "魔棒": 5000,            # 鹰身女妖之眼 85000 = 已知 80000 + 魔棒 50金
    "抗性合剂": 3000,        # 抗性强化药剂 10000 = 已知 70金 + 抗性合剂 30金
    "法杖": 15000,           # 储能魔杖 1275金 = 已知 1125金 + 法杖 150金
    "治愈因子": 3000,        # 医护机器人IR 225金 = 已知 75金 + 治愈因子5份 -> 30金/份
    "动力岩": 3000,          # 液压举重臂 85金 = 已知 25金 + 动力岩2份 -> 30金/份
    "黑火药": 2000,          # 爆炸陷阱 59金 = 已知 39金 + 黑火药1份 -> 20金/份
    "不稳定物质伽马": 90000,  # 世界放大器 1024金 = 已知 124金 + 伽马1份 -> 900金/份
    "辅助瞄准镜": 16000,      # 超光谱侦测镜 610金 = 已知 450金 + 瞄准镜1份 -> 160金/份
    "拆解的魔导枪": 18200,    # 奥术脉冲步枪 1177金 = 已知 995金 + 拆解魔导枪1份 -> 182金/份
    "一把霰弹枪": 5500,       # 瀑雨霰弹枪 2995金 = 已知 2940金 + 霰弹枪1份 -> 55金
    "强酸瓶": 500,           # 酸蚀炸弹 23.5金 = 已知 3.5金 + 强酸瓶2份 -> 5金/瓶
    "焦油": {"copper": 500, "weight_lb": 1, "per_lb": 500},  # 焦油陷阱 57金 = 已知 7金 + 10磅 -> 5金/磅
    "玩具乐器套装": 3000,     # 迷你乐团 45金 = 已知 15金 + 套装1份 -> 30金/份
    "彩色透镜组": 5300,       # 场景转换轮 287金 = 已知 234金 + 透镜组1份 -> 53金/份
    "磁感应雷电": 3500,       # 小型奥术浮游炮 705金 = 已知 670金 + 雷电1份 -> 35金/份
    "蜂蜜": 100,             # 物资大全 1金币/2磅，图纸按 1份 = 1金
}

# “任意X / 别名” -> 物资大全条目名
ANY_MAP = {
    "任意步枪": "步枪",
    "任意弹药": "弹药（20发）",
    "任意单手剑": "短剑",
    "任意单手枪械武器": "火枪",
    "任意一把单手枪械武器": "火枪",
    "一把单手枪械武器": "火枪",
    "任意步枪或渔枪": "步枪",
    "一把步枪或渔枪": "步枪",
    "任意绳": "麻绳（10米）",
    "任意锁链": "锁",
    "任意刀": "匕首",
    "任意皮甲": "皮甲",
    "任意掷矛": "标枪",
    "任意手枪": "火枪",
    "任意盾牌": "木质盾牌",
    "任意拳套": "拳刃",
    "一对拳套": "拳刃",
    "任意三把匕首": "匕首",
    "任意魔棒": "魔棒",
    "任意抗性合剂": "抗性合剂",
    "任意法杖": "法杖",
}
ALIAS = {
    "麻绳": "麻绳（10米）",
    "秘银": "秘银锭",
    "泡沫塑料": "塑料泡沫",
    "跃迁兽皮革": "跃迁兽的皮革",
}


def _extra_entry(v):
    if isinstance(v, dict):
        return dict(v)
    return {"copper": v, "weight_lb": None, "per_lb": None}


def lookup_price(name: str, D: dict) -> dict | None:
    name = name.strip()
    if name in D:
        return D[name]
    if name in ALIAS and ALIAS[name] in D:
        return D[ALIAS[name]]
    if name in ANY_MAP:
        target = ANY_MAP[name]
        if target in D:
            return D[target]
        if target in EXTRA_PRICE:
            return _extra_entry(EXTRA_PRICE[target])
    if name in EXTRA_PRICE:
        return _extra_entry(EXTRA_PRICE[name])
    m = re.match(r"^(.+?)[（(]", name)
    if m and m.group(1) in D:
        return D[m.group(1)]
    m2 = re.match(r"^(.+?)的(.+)$", name)
    if m2:
        alt = m2.group(1) + m2.group(2)
        if alt in D:
            return D[alt]
    return None


# ---------------------------------------------------------------------------
# 材料解析
# ---------------------------------------------------------------------------
QTY_RE = re.compile(
    r"(\d+(?:/\d+)?)\s*(份|磅|米|发|支|颗|个|把|张|组|块|根|件|对|枚|次|株|台|套|片|瓶|卷|袋|盏)$"
)


def parse_materials(line: str):
    """返回 [(name, qty_text, unit, raw_quantity)]。"""
    line = (line or "").strip()
    if not line or line == "-":
        return []
    body = line.split("：", 1)[-1]
    out = []
    for part in body.split("+"):
        part = part.strip()
        if not part:
            continue
        qm = QTY_RE.search(part)
        if qm:
            qty, unit = qm.group(1), qm.group(2)
            name = part[: qm.start()].strip()
            if "、" in name and "宝石" in name:
                for sub in name.replace("各", "").split("、"):
                    sub = sub.strip()
                    if not sub.endswith("宝石"):
                        sub = sub + "宝石"
                    out.append((sub, qty, unit, float_qty(qty)))
            else:
                out.append((name, qty, unit, float_qty(qty)))
        else:
            out.append((part, None, None, 1.0))
    return out


def float_qty(q):
    if not q:
        return 1.0
    if "/" in q:
        a, b = q.split("/")
        return float(a) / float(b)
    return float(q)


def fmt_qty(qty, unit):
    if not qty:
        return ""
    return qty + (unit or "")


# ---------------------------------------------------------------------------
# 计算
# ---------------------------------------------------------------------------
def compute_cost(items, D, get_plan, memo, stack, official):
    """返回 (known_copper:int, missing:[str])。get_plan(name)->dict|None 取成品计划。"""
    known = 0
    missing: list[str] = []
    for name, qty, unit, qn in items:
        ent = lookup_price(name, D)
        if ent is not None:
            if unit == "磅":
                if ent["per_lb"] is None:
                    missing.append(name + fmt_qty(qty, unit))
                    continue
                known += int(round(ent["per_lb"] * qn))
            elif unit == "米":
                known += int(round(ent["copper"] / 10.0 * qn))
            else:
                known += int(round(ent["copper"] * qn))
            continue
        # 官方有价成品（如魔导枪 182金）直接按官方价计
        off = None
        for cand in (name, name + "（图纸）", name.replace("（图纸）", "")):
            if cand in official:
                off = official[cand]
                break
        if off is not None:
            known += int(round(off * qn))
            continue
        # 图纸成品依赖（本次缺价计划内）
        plan = None
        for cand in (name, name + "（图纸）", name.replace("（图纸）", "")):
            p = get_plan(cand)
            if p is not None:
                plan = p
                break
        if plan is not None:
            if name in stack:
                missing.append(name + fmt_qty(qty, unit))
                continue
            pk, pm = compute_cost(
                plan["items"], D, get_plan, memo, stack + [name], official
            )
            known += int(round(pk * qn))
            for m in pm:
                tag = m + (fmt_qty(qty, unit) if qn != 1 else "")
                if tag not in missing:
                    missing.append(tag)
            continue
        tag = name + fmt_qty(qty, unit)
        if tag not in missing:
            missing.append(tag)
    return known, missing


def price_text(known_copper: int, missing: list[str]) -> str:
    known_copper = int(round(known_copper))
    silver = (known_copper + 9) // 10  # 向上取整到银币
    g, s = divmod(silver, 10)
    parts = []
    if g:
        parts.append(f"{g}金币")
    if s:
        parts.append(f"{s}银币")
    if not parts:
        parts.append("0银币")
    text = "参考价格：" + "".join(parts)
    if missing:
        text += "+" + "+".join(missing)
    return text


def close_existing_price_rows(skills: list[dict], D: dict) -> list[tuple]:
    """闭合 JSON 中已存在的“参考价格：X金币+材料”行（材料价全部已知时替换为完整价）。"""
    changed: list[tuple] = []
    for sk in skills:
        desc = sk.get("description") or []
        new_desc: list[str] = []
        for line in desc:
            m = re.match(r"^参考价格：(.+)$", line)
            if not m or "+" not in m.group(1):
                new_desc.append(line)
                continue
            pm = re.match(r"^(.*?)(\+\+?.*)$", m.group(1))
            if not pm:
                new_desc.append(line)
                continue
            base = parse_price_copper(pm.group(1))
            if base is None:
                new_desc.append(line)
                continue
            items = parse_materials("材料：" + pm.group(2).lstrip("+"))
            total = base
            ok = True
            for nm, qty, unit, qn in items:
                ent = lookup_price(nm, D)
                if ent is None:
                    ok = False
                    break
                if unit == "磅":
                    if ent.get("per_lb") is None:
                        ok = False
                        break
                    total += int(round(ent["per_lb"] * qn))
                elif unit == "米":
                    total += int(round(ent["copper"] / 10.0 * qn))
                else:
                    total += int(round(ent["copper"] * qn))
            if ok:
                new_line = price_text(total, [])
                new_desc.append(new_line)
                changed.append((sk.get("name", "?"), line, new_line))
            else:
                new_desc.append(line)
        if new_desc != desc:
            sk["description"] = new_desc
    return changed


# ---------------------------------------------------------------------------
# docx 提取
# ---------------------------------------------------------------------------
def official_prices(tables) -> dict[str, int]:
    """docx 中已有官方参考价的图纸：name -> copper。"""
    out = {}
    for t in tables:
        txt = "\n".join(r.cells[0].text for r in t.rows)
        if "研发材料" not in txt or "参考价格" not in txt:
            continue
        name = table_first_name(t)
        if not name:
            continue
        for r in t.rows:
            c = r.cells[0].text.strip()
            if c.startswith("参考价格"):
                p = parse_price_copper(c.split("：", 1)[-1])
                if p is not None:
                    out[name] = p
                break
    return out


def docx_tables(path: Path):
    d = docx.Document(str(path))
    return list(d.tables)


def table_first_name(t):
    cells = [r.cells[0].text.strip() for r in t.rows]
    return cells[0].split("\n")[0].strip() if cells else ""


def table_material_line(t, prefix="研发材料"):
    for r in t.rows:
        c = r.cells[0].text.strip()
        if c.startswith(prefix):
            return c
    return ""


def main() -> None:
    apply = "--apply" in sys.argv
    D = build_price_dict()

    # ---------- 1. 奇械师：从 docx 提取缺价条目 ----------
    art_tables = docx_tables(ROOT / "基础职业-奇械师.docx")
    qj_json = json.loads((ROOT / "职业页/数据/奇械师.json").read_text(encoding="utf-8"))
    qj_by_name = {s["name"]: s for s in qj_json["skills"]}

    # 缺价“研发材料”条目（docx 无参考价格）
    missing_skill_lines: dict[str, str] = {}
    # 加工行归属：最近的前一个技能表
    process_rows: list[dict] = []
    last_skill = None
    for t in art_tables:
        txt = "\n".join(r.cells[0].text for r in t.rows)
        name = table_first_name(t)
        is_skill_like = (
            "研发材料" in txt or "研发时间" in txt or "关键词" in txt or "参考价格" in txt
        )
        if "加工材料" in txt:
            mat = table_material_line(t, "加工材料")
            process_rows.append({"owner": last_skill, "mat": mat})
            continue
        if is_skill_like and name and "加工材料" not in name:
            last_skill = name
            if "研发材料" in txt and "参考价格" not in txt:
                mat = table_material_line(t, "研发材料")
                if mat:
                    missing_skill_lines[name] = mat

    # 网页缺“研发材料”行但 docx 缺价的技能
    web_missing_mat: dict[str, str] = {}
    for name, mat in missing_skill_lines.items():
        sk = qj_by_name.get(name)
        if sk is None:
            continue
        desc = sk.get("description") or []
        if not any(x.startswith("研发材料") for x in desc):
            web_missing_mat[name] = mat

    # 需补价的研发材料技能（网页 desc 已有研发材料行）
    web_price_skills: dict[str, str] = {}
    for name, mat in missing_skill_lines.items():
        sk = qj_by_name.get(name)
        if sk is None:
            continue
        desc = sk.get("description") or []
        if any(x.startswith("研发材料") for x in desc) and not any(
            x.startswith("参考价格") for x in desc
        ):
            web_price_skills[name] = mat

    # ---------- 2. 艺术与创造之神加工行 ----------
    art_docx = docx_tables(ROOT / "牧师子分支/神圣领域-艺术与创造之神.docx")
    art_process: list[dict] = []
    last_art = None
    for t in art_docx:
        txt = "\n".join(r.cells[0].text for r in t.rows)
        name = table_first_name(t)
        if "加工材料" in txt:
            art_process.append({"owner": last_art, "mat": table_material_line(t, "加工材料")})
            continue
        is_skill_like = (
            "研发材料" in txt or "研发时间" in txt or "参考价格" in txt
        )
        if is_skill_like and name and name != "类别" and "加工材料" not in name:
            last_art = name

    # ---------- 3. 生成计划条目 ----------
    official = official_prices(art_tables)
    plan: dict[str, dict] = {}

    def add_plan(key, owner, kind, mat_line, insert_mat=False):
        items = parse_materials(mat_line)
        plan[key] = {
            "owner": owner,
            "kind": kind,
            "mat": mat_line,
            "items": items,
            "insert_mat": insert_mat,
        }

    for name, mat in web_price_skills.items():
        if mat.split("：", 1)[-1].strip() == "-":
            continue  # 疫苗射线等无材料 -> 特例待定
        add_plan(name, name, "skill", mat, insert_mat=False)
    for name, mat in web_missing_mat.items():
        if mat.split("：", 1)[-1].strip() == "-":
            continue
        add_plan(name, name, "skill+mat", mat, insert_mat=True)
    for idx, pr in enumerate(process_rows):
        if pr["owner"]:
            add_plan(f"proc-qj-{idx}", pr["owner"], "process", pr["mat"], insert_mat=True)
    for idx, pr in enumerate(art_process):
        if pr["owner"]:
            add_plan(f"proc-art-{idx}", pr["owner"], "process", pr["mat"], insert_mat=True)

    def get_plan(name):
        return plan.get(name)

    memo: dict[str, tuple] = {}
    for key, p in plan.items():
        known, missing = compute_cost(
            p["items"], D, get_plan, memo, [key], official
        )
        p["known"] = known
        p["missing"] = missing
        p["price"] = price_text(known, missing)
        memo[key] = (known, missing)

    out = {
        "price_dict_size": len(D),
        "plans": plan,
        "special": {
            "疫苗射线": "待定（研发材料：-）",
            "全感知乐谱（图纸）": "待定（材料为空）",
        },
    }
    out_path = ROOT / "scripts/_price_plan.json"
    out_path.write_text(
        json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8"
    )

    # 打印摘要
    print(f"价格字典条目: {len(D)}")
    print(f"计划条目: {len(plan)}")
    stats = {"skill": 0, "skill+mat": 0, "process": 0}
    for p in plan.values():
        stats[p["kind"]] += 1
    print("  研发材料缺价:", stats["skill"], "| 研发材料行缺失:", stats["skill+mat"], "| 加工行:", stats["process"])
    owners = sorted({p["owner"] for p in plan.values() if p["kind"] == "process"})
    print("  加工行归属技能数:", len(owners), owners)
    all_missing: dict[str, int] = {}
    for p in plan.values():
        for m in p["missing"]:
            base = re.sub(r"[\d/]+(?:份|磅|米|株|台|套|片|瓶|卷|袋|盏|发|支|颗|个|把|张|组|块|根|件|对|枚|次)$", "", m)
            all_missing[base] = all_missing.get(base, 0) + 1
    print("缺失材料种类:", len(all_missing))
    for m, c in sorted(all_missing.items(), key=lambda x: -x[1]):
        print(f"   {m}: {c}")
    print()
    for key in sorted(plan, key=lambda k: (plan[k]["owner"], k)):
        p = plan[key]
        print(f"  [{p['kind']}] {p['owner'][:22]:24} -> {p['price']}")
    print()
    print("特例:", out["special"])

    if not apply:
        print("\n[dry] 未写入任何文件。")
        return

    # ---------- 4. 写入 JSON ----------
    for key, p in plan.items():
        owner = p["owner"]
        sk = qj_by_name.get(owner)
        if sk is None:
            print("  !! 归属技能缺失:", owner)
            continue
        desc = sk.get("description") or []
        new_desc = list(desc)
        if p["kind"] in ("skill", "skill+mat"):
            if p["insert_mat"]:
                # 补“研发材料”行（放在研发时间前）
                idx = next((i for i, x in enumerate(new_desc) if x.startswith("研发时间")), 0)
                new_desc.insert(idx, p["mat"])
            if not any(x.startswith("参考价格") for x in new_desc):
                price_idx = next(
                    (i for i, x in enumerate(new_desc) if x.startswith("研发时间")),
                    len(new_desc),
                )
                new_desc.insert(price_idx + 1, p["price"])
        else:  # process
            if not any(x.startswith("加工材料") and p["mat"].split("：", 1)[-1] in x for x in new_desc):
                new_desc.append(p["mat"])
                new_desc.append(p["price"])
        sk["description"] = new_desc

    changed_qj = close_existing_price_rows(qj_json["skills"], D)
    for nm, old_line, new_line in changed_qj:
        print(f"  闭合: {nm} | {old_line} -> {new_line}")

    qj_json_path = ROOT / "职业页/数据/奇械师.json"
    qj_json_path.write_text(
        json.dumps(qj_json, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print("奇械师.json 已写入")

    # 艺术与创造之神（牧师·神圣领域.json）
    dom_json = json.loads(
        (ROOT / "职业页/数据/牧师·神圣领域.json").read_text(encoding="utf-8")
    )
    art_dom = next(
        (d for d in dom_json["domains"].values() if d["name"] == "艺术与创造之神"), None
    )
    if art_dom:
        by_name = {s["name"]: s for s in art_dom["skills"]}
        for key, p in plan.items():
            if not key.startswith("proc-art"):
                continue
            sk = by_name.get(p["owner"]) or by_name.get(p["owner"] + "（图纸）")
            if sk is None:
                print("  !! 艺术归属技能缺失:", p["owner"])
                continue
            desc = sk.get("description") or []
            new_desc = list(desc)
            if not any(
                x.startswith("加工材料") and p["mat"].split("：", 1)[-1] in x
                for x in new_desc
            ):
                new_desc.append(p["mat"])
                new_desc.append(p["price"])
            sk["description"] = new_desc
        changed_dom = close_existing_price_rows(art_dom["skills"], D)
        for nm, old_line, new_line in changed_dom:
            print(f"  闭合: {nm} | {old_line} -> {new_line}")

        dom_json_path = ROOT / "职业页/数据/牧师·神圣领域.json"
        dom_json_path.write_text(
            json.dumps(dom_json, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print("牧师·神圣领域.json 已写入")

    # ---------- 5. 重建受影响的 HTML 卡片 ----------
    affected = {}
    for key, p in plan.items():
        if not key.startswith("proc-art"):
            affected.setdefault("奇械师", set()).add(p["owner"])
        else:
            affected.setdefault("牧师", set()).add(p["owner"])
    affected.setdefault("奇械师", set()).update(nm for nm, _, _ in changed_qj)
    affected.setdefault("牧师", set()).update(nm for nm, _, _ in changed_dom)

    for cls, names in affected.items():
        if cls == "奇械师":
            data = qj_json
            html_path = ROOT / "职业页/奇械师.html"
        else:
            data = dom_json
            html_path = ROOT / "职业页/牧师.html"
        html = html_path.read_text(encoding="utf-8")
        if cls == "奇械师":
            skills = data["skills"]
        else:
            skills = art_dom["skills"]
        for sk in skills:
            if sk.get("name") not in names:
                continue
            sid = sk["id"]
            if f'id="{sid}"' not in html:
                print("  !! 页面未找到:", sid, sk.get("name"))
                continue
            block = extract_to_block(sk)
            mark_dots = []
            for c in block.get("mark_dots") or []:
                if isinstance(c, dict):
                    mark_dots.extend([c["color"]] * c.get("count", 1))
                else:
                    mark_dots.append(c)
            block["mark_dots"] = mark_dots
            tables = {
                "unit_tables": sk.get("unit_tables") or [],
                "roll_tables": sk.get("roll_tables") or [],
            }
            detail = build_detail_html(block, tables)
            tier_lbl = tier_label_from_skill(sk)
            data_search = build_data_search(
                block, sk.get("style") or "", tier_lbl, sk.get("tags") or []
            )
            data_search = append_tables_to_search(data_search, sk)
            data_attrs = build_skill_data_attrs(sk, block.get("mark_dots") or [], cls)
            html = patch_html(html, sid, detail, data_search, data_attrs)
            if cls == "牧师":
                deity = next(
                    (
                        d["name"]
                        for d in dom_json["domains"].values()
                        if sk in d.get("skills", [])
                    ),
                    "",
                )
                if deity:
                    html = html.replace(
                        f'<article class="skill" id="{sid}"',
                        f'<article class="skill" id="{sid}" data-deity="{deity}"',
                        1,
                    )
        html_path.write_text(html, encoding="utf-8")
        print(f"{cls}.html 已重建 ({len(names)} 张卡片)")

    # ---------- 6. 重建 skill_effects（奇械师） ----------
    fx_path = ROOT / "斯诺德跑团/skill_effects_奇械师.json"
    fx = {"奇械师": [json_to_fx_entry(s, "奇械师") for s in qj_json["skills"]]}
    fx_path.write_text(json.dumps(fx, ensure_ascii=False, indent=2), encoding="utf-8")
    print("skill_effects_奇械师.json 已重建")

    # ---------- 7. 镜像 electron-app + 搜索索引 ----------
    for rel in [
        "职业页/奇械师.html",
        "职业页/数据/奇械师.json",
        "斯诺德跑团/skill_effects_奇械师.json",
        "职业页/牧师.html",
        "职业页/数据/牧师·神圣领域.json",
    ]:
        dst = ROOT / "electron-app" / rel
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(ROOT / rel, dst)
    subprocess.check_call(
        ["node", str(ROOT / "scripts/build_class_search_index.js")],
        cwd=str(ROOT),
    )
    shutil.copy2(
        ROOT / "职业页/search-index.json",
        ROOT / "electron-app/职业页/search-index.json",
    )
    shutil.copy2(
        ROOT / "职业页/search-index.js",
        ROOT / "electron-app/职业页/search-index.js",
    )
    print("镜像与搜索索引已更新")


if __name__ == "__main__":
    main()
