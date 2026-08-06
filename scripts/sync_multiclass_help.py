# -*- coding: utf-8 -*-
"""
从 冒险者基础规则.xlsx「兼职规则」工作表同步帮助页 s4 区块。

生成内容：
1. 兼职规则表（职业 / 属性要求 / 熟练度要求 / 其他要求 / 不可兼职）
2. 职业兼容性稀疏矩阵（仅标红不可兼职，空白=可兼职，灰 —=自身）
3. 配套 CSS（粘性首列/表头、悬停高亮、图例）与 hover JS

用法：
    python scripts/sync_multiclass_help.py           # 同步并校验
    python scripts/sync_multiclass_help.py --verify  # 仅校验
"""
from __future__ import annotations

import argparse
import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "冒险者基础规则.xlsx"
HELP = ROOT / "斯诺德跑团" / "help.html"
ELECTRON_HELP = ROOT / "electron-app" / "斯诺德跑团" / "help.html"

CLASSES = [
    "蛮斗士", "战士", "法师", "猎人", "牧师", "圣骑士", "游荡者",
    "德鲁伊", "萨满祭司", "术士", "武僧", "吟游诗人", "魔契师", "奇械师",
]

# 矩阵列头缩写（title 显示全名）
SHORT = {
    "蛮斗士": "蛮斗", "战士": "战士", "法师": "法师", "猎人": "猎人",
    "牧师": "牧师", "圣骑士": "圣骑", "游荡者": "游荡", "德鲁伊": "德鲁",
    "萨满祭司": "萨满", "术士": "术士", "武僧": "武僧", "吟游诗人": "吟游",
    "魔契师": "魔契", "奇械师": "奇械",
}

BEGIN_BODY = "<!-- MULTICLASS-RULES -->"
END_BODY = "<!-- /MULTICLASS-RULES -->"
BEGIN_CSS = "/* MULTICLASS-RULES-CSS */"
END_CSS = "/* /MULTICLASS-RULES-CSS */"


def load_xlsx() -> dict:
    """读取兼职规则工作表，返回 {职业: {attr, prof, other, incompatible:[]}}。"""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["兼职规则"]
    data = {}
    for r in range(3, 17):
        name = ws.cell(r, 2).value
        if not name:
            continue
        compatible = [
            ws.cell(r, c).value
            for c in range(6, 20)
            if ws.cell(r, c).value and ws.cell(r, c).value != "-"
        ]
        incompatible = [c for c in CLASSES if c not in compatible]
        data[name] = {
            "attr": str(ws.cell(r, 3).value or "").strip(),
            "prof": str(ws.cell(r, 4).value or "").strip(),
            "other": str(ws.cell(r, 5).value or "").strip(),
            "incompatible": incompatible,
        }
    return data


CSS = f"""
{BEGIN_CSS}
.mc-wrap{{overflow:auto;margin:10px 0;max-width:100%;max-height:72vh}}
.mc-matrix{{border-collapse:collapse;font-size:12px;white-space:nowrap}}
.mc-matrix th,.mc-matrix td{{border:1px solid #d8d2c4;padding:6px 9px;text-align:center;min-width:2.8em}}
.mc-matrix thead th{{position:sticky;top:0;z-index:2;background:#edeae2;color:#1f2522;font-weight:bold}}
.mc-matrix tbody th{{position:sticky;left:0;z-index:1;background:#edeae2;color:#1f2522;font-weight:bold;text-align:left}}
.mc-matrix thead th:first-child{{left:0;z-index:3}}
.mc-matrix td{{background:#fffdf8}}
.mc-matrix td.mc-no{{background:#c62828;color:#fff;font-weight:bold}}
.mc-matrix td.mc-self{{background:#e8e4dc;color:#69706b}}
.mc-matrix tr.mc-hl td{{background:#f3efe6}}
.mc-matrix tr.mc-hl td.mc-no{{background:#d94848}}
.mc-matrix td.mc-hl-col{{background:#f3efe6}}
.mc-matrix td.mc-hl-col.mc-no{{background:#d94848}}
.mc-matrix td.mc-hl-col.mc-self{{background:#e0dcd2}}
.mc-legend{{font-size:12px;color:#69706b;margin:6px 0 4px;display:flex;flex-wrap:wrap;gap:14px;align-items:center}}
.mc-legend .sw{{display:inline-flex;align-items:center;gap:5px}}
.mc-legend .sw::before{{content:"";display:inline-block;width:14px;height:14px;border:1px solid #d8d2c4;border-radius:3px;flex:0 0 auto}}
.mc-legend .sw.mc-no::before{{background:#c62828}}
.mc-legend .sw.mc-self::before{{background:#e8e4dc}}
.mc-hint{{font-size:12px;color:#69706b;margin:2px 0 8px}}
html.dark .mc-matrix thead th,html.dark .mc-matrix tbody th{{background:#1a1d20;color:#e8e6e3;border-color:#3a3d40}}
html.dark .mc-matrix td{{background:#24272b;border-color:#3a3d40;color:#e8e6e3}}
html.dark .mc-matrix td.mc-self{{background:#2c2f33;color:#9d9b98}}
html.dark .mc-matrix td.mc-no{{background:#a52a2a}}
html.dark .mc-matrix tr.mc-hl td,html.dark .mc-matrix td.mc-hl-col{{background:#33373b}}
html.dark .mc-legend,html.dark .mc-hint{{color:#9d9b98}}
{END_CSS}
"""


JS = """<script>
(function(){
  var t = document.querySelector('.mc-matrix');
  if (!t) return;
  function clear(){
    var n = t.querySelectorAll('.mc-hl-row, .mc-hl-col');
    for (var i = 0; i < n.length; i++) n[i].classList.remove('mc-hl-row', 'mc-hl-col');
  }
  t.addEventListener('mouseover', function(e){
    var td = e.target && e.target.closest ? e.target.closest('td,th') : null;
    if (!td || !t.contains(td)) return;
    clear();
    var tr = td.parentNode;
    if (tr && tr.rowIndex > 0) tr.classList.add('mc-hl-row');
    var ci = td.cellIndex;
    var rows = t.rows;
    for (var i = 0; i < rows.length; i++) {
      var c = rows[i].cells[ci];
      if (c) c.classList.add('mc-hl-col');
    }
  });
  t.addEventListener('mouseleave', clear);
})();
</script>"""


def build_rule_table(data: dict) -> str:
    rows = ['<div class="wrap"><table><tr><th>职业</th><th>属性要求</th><th>熟练度要求</th><th>其他要求</th><th>不可兼职</th></tr>']
    for name in CLASSES:
        d = data[name]
        inc = "、".join(d["incompatible"]) if d["incompatible"] else "—"
        rows.append(
            f'<tr><td><b>{name}</b></td><td>{d["attr"]}</td><td>{d["prof"]}</td>'
            f'<td>{d["other"]}</td><td>{inc}</td></tr>'
        )
    rows.append("</table></div>")
    return "\n".join(rows)


def build_matrix(data: dict) -> str:
    head = ["<tr><th>主\\副</th>"]
    for c in CLASSES:
        head.append(f'<th title="{c}">{SHORT[c]}</th>')
    head.append("</tr>")
    body = []
    for rn in CLASSES:
        cells = [f'<th title="{rn}">{rn}</th>']
        for cn in CLASSES:
            if rn == cn:
                cells.append(f'<td class="mc-self" title="{rn}（自身）">—</td>')
            elif cn in data[rn]["incompatible"]:
                cells.append(f'<td class="mc-no" title="{rn} 不能兼职 {cn}">✗</td>')
            else:
                cells.append(f'<td class="mc-ok" title="{rn} 可兼职 {cn}"></td>')
        body.append("<tr>" + "".join(cells) + "</tr>")
    return (
        '<h3>职业兼容性</h3>\n'
        '<div class="mc-legend">'
        '<span class="sw mc-no">不可兼职</span>'
        '<span class="sw mc-ok">可兼职（空白）</span>'
        '<span class="sw mc-self">自身</span>'
        '<span>行 = 主职业，列 = 副职</span>'
        '</div>\n'
        '<div class="mc-wrap"><table class="mc-matrix">'
        "<thead>" + "".join(head) + "</thead><tbody>"
        + "".join(body) + "</tbody></table></div>\n"
        + JS
    )


def build_block(data: dict) -> str:
    return (
        BEGIN_BODY
        + "\n"
        + build_rule_table(data)
        + "\n\n"
        + build_matrix(data)
        + "\n"
        + END_BODY
    )


def sync_css(html: str) -> str:
    if BEGIN_CSS in html and END_CSS in html:
        start = html.index(BEGIN_CSS)
        end = html.index(END_CSS) + len(END_CSS)
        return html[:start] + CSS + html[end:]
    marker = "</style>"
    assert marker in html, "未找到 </style>"
    return html.replace(marker, CSS + "\n" + marker, 1)


def sync_body(html: str, data: dict) -> str:
    block = build_block(data)
    if BEGIN_BODY in html and END_BODY in html:
        start = html.index(BEGIN_BODY)
        end = html.index(END_BODY) + len(END_BODY)
        return html[:start] + block + html[end:]

    # 首次：定位 s4 兼职规则区块，从 <h2>兼职规则</h2> 之后到附赠职业 note 之前整体替换
    anchor = '<div class="section" id="s4"><h2>兼职规则</h2>'
    assert anchor in html, "未找到 s4 兼职规则区块"
    start = html.index(anchor) + len(anchor)
    note = '<div class="note"><b>附赠职业：</b>'
    assert note in html[start:], "未找到附赠职业 note"
    end = html.index(note, start)
    return html[:start] + "\n" + block + "\n" + html[end:]


def write_preserve(path: Path, content: str) -> None:
    path.write_text(content, encoding="utf-8", newline="")


def read_preserve(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def verify(html: str, data: dict) -> list:
    import re

    errors = []
    body = html
    if BEGIN_BODY in body and END_BODY in body:
        body = body[body.index(BEGIN_BODY): body.index(END_BODY)]

    # 1) 兼职规则表：14 职业 × 4 列
    rows = re.findall(
        r"<tr><td><b>([^<]+)</b></td><td>([^<]*)</td><td>([^<]*)</td><td>([^<]*)</td><td>([^<]*)</td></tr>",
        body,
    )
    if len(rows) != 14:
        errors.append(f"兼职规则表行数 {len(rows)} != 14")
    for name, attr, prof, other, inc in rows:
        d = data.get(name)
        if not d:
            errors.append(f"表内未知职业: {name}")
            continue
        expect_inc = "、".join(d["incompatible"]) if d["incompatible"] else "—"
        if attr != d["attr"]:
            errors.append(f"{name} 属性: {attr} != {d['attr']}")
        if prof != d["prof"]:
            errors.append(f"{name} 熟练: {prof} != {d['prof']}")
        if other != d["other"]:
            errors.append(f"{name} 其他: {other} != {d['other']}")
        if inc != expect_inc:
            errors.append(f"{name} 不可兼职: {inc} != {expect_inc}")

    # 2) 稀疏矩阵：182 个方向对
    m = re.search(r'<table class="mc-matrix">.*?</table>', body, re.S)
    if not m:
        errors.append("未找到 mc-matrix 矩阵")
    else:
        cells = re.findall(r'<td class="(mc-no|mc-ok|mc-self)"[^>]*>', m.group(0))
        # 需要按行分组验证；直接按类顺序逐格解析
        matrix = re.findall(r"<tr>(.*?)</tr>", m.group(0), re.S)
        data_rows = matrix[1:]  # 去掉表头
        if len(data_rows) != 14:
            errors.append(f"矩阵数据行数 {len(data_rows)} != 14")
        for ri, rn in enumerate(CLASSES):
            if ri >= len(data_rows):
                break
            tds = re.findall(r'<td class="(mc-no|mc-ok|mc-self)"', data_rows[ri])
            if len(tds) != 14:
                errors.append(f"{rn} 矩阵列数 {len(tds)} != 14")
                continue
            for ci, cn in enumerate(CLASSES):
                expect = (
                    "mc-self" if rn == cn
                    else "mc-no" if cn in data[rn]["incompatible"]
                    else "mc-ok"
                )
                if tds[ci] != expect:
                    errors.append(f"矩阵 {rn}->{cn}: {tds[ci]} != {expect}")

    # 3) 旧矩阵残留检查
    if "color:#2e7d32" in body or "color:#c62828" in body:
        errors.append("s4 区块仍残留旧 ✓/✗ 矩阵的内联颜色")
    if "职业兼容性" not in body:
        errors.append("s4 区块缺少 职业兼容性 标题")
    return errors


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--verify", action="store_true", help="仅校验帮助页与 xlsx 是否一致")
    args = parser.parse_args()

    data = load_xlsx()
    html = read_preserve(HELP)
    errors = verify(html, data)

    if args.verify:
        if errors:
            print("VERIFY FAIL")
            for e in errors:
                print("  -", e)
            return 1
        print("VERIFY PASS：兼职规则表 56 项断言 + 矩阵 182 方向对全部一致")
        return 0

    html = sync_css(html)
    html = sync_body(html, data)
    write_preserve(HELP, html)
    shutil.copyfile(HELP, ELECTRON_HELP)

    errors = verify(read_preserve(HELP), data)
    if errors:
        print("SYNC FAIL")
        for e in errors:
            print("  -", e)
        return 1
    print("SYNC OK：help.html 已更新并同步 electron-app")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
